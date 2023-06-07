import openpyxl
import sys
import os
from PySide6 import QtUiTools, QtGui
from PySide6.QtWidgets import QApplication, QMainWindow
from PySide6.QtGui import *
from PySide6.QtCore import *

class MainView(QMainWindow):    
    def __init__(self):
        super().__init__()
        self.setupUI()

    def setupUI(self):
        global UI_set

        UI_set = QtUiTools.QUiLoader().load(resource_path("wrp_start.ui"))
        self.setCentralWidget(UI_set)
        self.setWindowTitle("UI TEST")
        self.setWindowIcon(QtGui.QPixmap(resource_path("./images/word.png")))
        self.resize(270,500)
        self.show()

        UI_set.start.clicked.connect(self.Main)

    def Main (self):
        global Main

        Main = QtUiTools.QUiLoader().load(resource_path("wrp_main.ui"))
        self.setCentralWidget(Main)
        self.setWindowTitle("List")
        self.setWindowIcon(QtGui.QPixmap(resource_path("./images/word.png")))
        self.resize(270,500)
        self.show()

        list = openpyxl.load_workbook('wordfile.xlsx')
        name = list.get_sheet_names()
        for i in name:
            Main.wordList.addItem(i)
        
        Main.wordList.itemDoubleClicked.connect(self.Word)
        Main.deleteButton.clicked.connect(self.delete)
        Main.viewButton.clicked.connect(self.List)

    def delete (self):
        list = openpyxl.load_workbook('wordfile.xlsx')
        name = list.get_sheet_names()
        chap = Main.wordList.currentRow()
        print(name[chap])
        list.remove_sheet(list.get_sheet_by_name(name[chap]))
        list.save('wordfile.xlsx')
        Main.wordList.takeItem(chap)
        self.Main()

    #def wordplace (self,name):
        #chap = Main.wordList.currentRow()
        #self.Word(chap)
    i=0
    def Word (self):
        global Word

        Word = QtUiTools.QUiLoader().load(resource_path("wrp_word.ui"))
        self.setCentralWidget(Word)
        self.setWindowTitle("Word")
        self.setWindowIcon(QtGui.QPixmap(resource_path("./images/word.png")))
        self.resize(270,500)
        self.show()

        list = openpyxl.load_workbook('wordfile.xlsx')
        name = list.get_sheet_names()
        chap = Main.wordList.currentRow()
        Word.chaptername.setText(name[chap])

        sheet = list.get_sheet_by_name(name[chap])
        count=0
        for row in sheet:
            count +=1
        print(count)
            
        global W
        W=[]
        for row in sheet:
            W.append(row[0].value)
        
        global M
        M=[]
        for row in sheet:
            M.append(row[1].value)

        self.i=0
        if self.i < count:
            Word.nextButton.clicked.connect(self.update)
            print(self.i)
        elif self.i==count-1:
            Word.word.setText("마지막 단어 입니다.")
            Word.means.setText("마지막 단어 입니다.")
        
        Word.endButton.clicked.connect(self.Main)
        
    
    def update (self):
        Word.word.setText(W[self.i])
        Word.means.setText(M[self.i])
        self.i +=1
        
    def List (self):
        global List

        List = QtUiTools.QUiLoader().load(resource_path("wrp_list.ui"))
        self.setCentralWidget(List)
        self.setWindowTitle("Word List")
        self.setWindowIcon(QtGui.QPixmap(resource_path("./images/word.png")))
        self.resize(270,500)
        self.show()
        
        list = openpyxl.load_workbook('wordfile.xlsx')
        name = list.get_sheet_names()
        chap = Main.wordList.currentRow()
        List.Chapter.setText(name[chap])

        sheet = list.get_sheet_by_name(name[chap])
        count=0
        for row in sheet:
            count +=1
        W=[]
        for row in sheet:
            W.append(row[0].value)
        
        model =QStandardItemModel()
        for i in W:
            model.appendRow(QStandardItem(i))
        List.wordView.setModel(model)

        M=[]
        for row in sheet:
            M.append(row[1].value)
        
        model =QStandardItemModel()
        for i in M:
            model.appendRow(QStandardItem(i))
        List.meanView.setModel(model)

        List.backButton.clicked.connect(self.Main)


#파일 경로
#pyinstaller로 원파일로 압축할때 경로 필요함
def resource_path(relative_path):
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)
 
if __name__ == '__main__':
    app = QApplication(sys.argv)
    main = MainView()
    main.show()
    sys.exit(app.exec())
